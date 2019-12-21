import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

current_dir = os.path.dirname(__file__)
flname = os.path.join(current_dir, "Book.xlsx")
#path="C:\Users\aanch\Desktop\wb.csv"

bk=openpyxl.load_workbook(flname)
anotherSheet = bk.active


sheet = bk['sheet no 1']
xl=[]
for i in range(65,91):
    xl.append(chr(i))
sheet.title='sheet no 1'

print(bk.sheetnames)
'''
wb.create_sheet(index=1,title='try')
sheet=wb['try']
'''
for i in range(1,101):
    for j in xl:
        sheet[j+ str(i)]=j+str(i)

bk.save(flname)




